import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "asksuite-market.json"

DDD_UF = {
    "11": "SP", "12": "SP", "13": "SP", "14": "SP", "15": "SP", "16": "SP", "17": "SP", "18": "SP", "19": "SP",
    "21": "RJ", "22": "RJ", "24": "RJ", "27": "ES", "28": "ES",
    "31": "MG", "32": "MG", "33": "MG", "34": "MG", "35": "MG", "37": "MG", "38": "MG",
    "41": "PR", "42": "PR", "43": "PR", "44": "PR", "45": "PR", "46": "PR",
    "47": "SC", "48": "SC", "49": "SC", "51": "RS", "53": "RS", "54": "RS", "55": "RS",
    "61": "DF", "62": "GO", "64": "GO", "63": "TO", "65": "MT", "66": "MT", "67": "MS",
    "68": "AC", "69": "RO", "71": "BA", "73": "BA", "74": "BA", "75": "BA", "77": "BA", "79": "SE",
    "81": "PE", "87": "PE", "82": "AL", "83": "PB", "84": "RN", "85": "CE", "88": "CE", "86": "PI", "89": "PI",
    "91": "PA", "93": "PA", "94": "PA", "92": "AM", "97": "AM", "95": "RR", "96": "AP", "98": "MA", "99": "MA",
}


def clean(value):
    return str(value or "").strip()


def number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("R$", "").replace("%", "").replace(" ", "").strip()
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def date_key(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = clean(value)
    iso = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    if iso:
        return f"{iso.group(1)}-{iso.group(2)}-{iso.group(3)}"
    br = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if br:
        return f"{br.group(3)}-{br.group(2).zfill(2)}-{br.group(1).zfill(2)}"
    return ""


def phone_ddd(value):
    digits = re.sub(r"\D+", "", clean(value))
    if digits.startswith("55") and len(digits) >= 4:
        return digits[2:4]
    if len(digits) >= 10:
        return digits[:2]
    return ""


def norm_header(value):
    text = clean(value).upper()
    replacements = str.maketrans("ÁÀÂÃÉÊÍÓÔÕÚÇ", "AAAAEEIOOOUC")
    return text.translate(replacements)


def canonical_hotel(value):
    key = norm_header(value)
    hotels = {
        "SUEDS HOTELS": "SUEDS Hotels",
        "SUEDS CABRALIA": "SUEDS Cabralia",
        "SUEDS PLAZA": "SUEDS Plaza",
        "SUEDS SEGUNDO SOL": "SUEDS Segundo Sol",
        "SUEDS PREMIUM": "SUEDS Premium",
        "SUEDS TRANCOSO": "SUEDS Trancoso",
        "CASAS SUEDS ARRAIAL": "Casas Sueds Arraial",
    }
    return hotels.get(key, clean(value) or "Não informado")


def canonical_channel(value):
    key = norm_header(value)
    if key == "WHATSAPP":
        return "WhatsApp"
    if key == "CHAT_WEB":
        return "Chat web"
    if key == "INSTAGRAM":
        return "Instagram"
    return clean(value) or "Não informado"


def is_robot_attendant(value):
    return norm_header(value) in {"ROBO", "ROBÔ"}


def parse_file(path, month):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [clean(cell.value) for cell in next(sheet.iter_rows(max_row=1))]
    indexes = {header: index for index, header in enumerate(headers)}
    required = ["Telefone", "Atendente", "Empresa", "Canal", "Início do atendimento", "Oportunidades", "Vendas", "Valor vendido"]
    missing = [header for header in required if header not in indexes]
    if missing:
        raise RuntimeError(f"{path.name}: colunas ausentes: {', '.join(missing)}")

    groups = defaultdict(lambda: {
        "dialogues": 0,
        "quotes": 0.0,
        "reservations": 0.0,
        "sales": 0.0,
        "revenue": 0.0,
    })

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        start = date_key(row[indexes["Início do atendimento"]])
        row_month = start[:7] if start else month
        if row_month != month:
            continue
        ddd = phone_ddd(row[indexes["Telefone"]])
        state = DDD_UF.get(ddd, "Não informado")
        hotel = canonical_hotel(row[indexes["Empresa"]])
        channel = "Robo" if is_robot_attendant(row[indexes["Atendente"]]) else canonical_channel(row[indexes["Canal"]])
        key = (month, state, ddd or "NI", hotel, channel)
        target = groups[key]
        opportunities = number(row[indexes["Oportunidades"]])
        target["dialogues"] += 1
        target["quotes"] += opportunities
        target["reservations"] += opportunities
        target["sales"] += number(row[indexes["Vendas"]])
        target["revenue"] += number(row[indexes["Valor vendido"]])

    rows = []
    for (month, state, ddd, hotel, channel), values in groups.items():
        rows.append({
            "month": month,
            "state": state,
            "ddd": ddd,
            "hotel": hotel,
            "channel": channel,
            "campaign": f"Asksuite {channel}",
            "source": "Asksuite",
            "origin": channel,
            "device": "Não informado",
            "dialogues": int(values["dialogues"]),
            "quotes": round(values["quotes"], 2),
            "reservations": round(values["reservations"], 2),
            "sales": round(values["sales"], 2),
            "revenue": round(values["revenue"], 2),
            "googleSpend": 0,
            "metaSpend": 0,
        })
    return rows


def main():
    if len(sys.argv) < 3:
        raise SystemExit("Uso: python tools/build_asksuite_market_data.py 2026-05=arquivo.xlsx 2026-06=arquivo.xlsx")

    all_rows = []
    sources = []
    for arg in sys.argv[1:]:
        if "=" not in arg:
            raise SystemExit(f"Argumento invalido: {arg}")
        month, file_name = arg.split("=", 1)
        path = Path(file_name)
        rows = parse_file(path, month)
        all_rows.extend(rows)
        sources.append({"month": month, "file": path.name, "rows": len(rows)})

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": "Asksuite Sales Report",
        "sources": sources,
        "rows": sorted(all_rows, key=lambda row: (row["month"], row["state"], row["ddd"], row["hotel"], row["channel"])),
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output": str(OUTPUT),
        "rows": len(payload["rows"]),
        "sources": sources,
        "months": sorted({row["month"] for row in payload["rows"]}),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
