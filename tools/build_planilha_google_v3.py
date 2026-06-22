from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\roger\OneDrive\Documentos\Dashboard SUEDS HOTELS")
INPUT = ROOT / "Vendas Sueds Hotels - V19jun26.xlsx"
OUTPUT = ROOT / "Vendas Sueds Hotels - V3 Google Sheets.xlsx"


COLORS = {
    "dark": "10201B",
    "header": "1F4E3D",
    "soft": "EAF4EF",
    "line": "D9E2EC",
    "warn": "FFF2CC",
    "input": "F7FBF8",
    "locked": "EEF2F6",
    "text": "FFFFFF",
}


PROTECTION_PASSWORD = "sueds"


ENTRY_HEADERS = [
    "Data Venda",
    "Codigo Reserva",
    "Hotel",
    "Canal",
    "Vendedor",
    "Cliente",
    "Checkin",
    "Checkout",
    "Diarias",
    "UHs",
    "Adultos",
    "Criancas",
    "Valor Total",
    "Recebido",
    "A Receber",
    "Forma Pagto",
    "Parcelas",
    "Status",
    "Fonte",
    "Observacoes",
]


BASE_HEADERS = ENTRY_HEADERS + [
    "Hotel Normalizado",
    "Canal Macro",
    "Canal Detalhado",
    "Tipo Responsavel",
    "Mes Venda",
]


def normalize_hotel(value: str) -> str:
    mapping = {
        "SEGUNDO SOL": "SUEDS SEGUNDO SOL",
        "SUEDS SEGUNDO SOL": "SUEDS SEGUNDO SOL",
        "SEGUNDO CABRALIA": "SUEDS CABRALIA",
        "SUEDS CABRALIA": "SUEDS CABRALIA",
        "SUEDS TRANCOSO": "SUEDS TRANCOSO",
        "SUEDS PLAZA": "SUEDS PLAZA",
        "SUEDS PREMIUM": "SUEDS PREMIUM",
    }
    return mapping.get((value or "").strip().upper(), value or "")


def channel_fields(channel: str, seller: str) -> tuple[str, str, str]:
    channel_key = (channel or "").strip().upper()
    seller_key = (seller or "").strip().upper()
    if channel_key in {"BOOKING ENGINE", "BE MOBILE", "SITE"}:
        return "Direto", "Booking Engine", "Canal digital"
    if channel_key in {"CENTRAL DE RESERVAS", "CENTRAL"}:
        return "Direto", "Central de Reservas", "Atendente"
    if channel_key in {"WHATSAPP", "PARTICULAR"}:
        return "Direto", "WhatsApp", "Atendente" if seller_key != "SITE" else "Canal digital"
    if channel_key == "OPERADORAS":
        return "Operadora", "Operadoras", "Parceiro"
    if channel_key == "OTAS":
        return "OTA", "OTAs", "Parceiro"
    return "Direto", channel or "Nao informado", "Atendente"


def payment_fields(payment: str) -> tuple[str, str]:
    text = (payment or "").upper()
    match = re.search(r"(\d+)\s*X", text)
    parcelas = match.group(1) if match else ""
    if "PIX" in text:
        return "PIX", parcelas
    if "CREDITO" in text or "CRÉDITO" in text or "CARTAO" in text or "CARTÃO" in text or "LINK" in text:
        return "Cartao credito", parcelas
    if "DEBITO" in text or "DÉBITO" in text:
        return "Cartao debito", parcelas
    if "DINHEIRO" in text:
        return "Dinheiro", parcelas
    if "VALOR MENCIONADO" in text:
        return "Ver linha anterior", parcelas
    return "Outros", parcelas


def style_header(ws, row: int, start_col: int, end_col: int, fill: str = COLORS["header"]) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=COLORS["text"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=COLORS["line"]))


def add_table(ws, name: str, first_row: int, first_col: int, last_row: int, last_col: int) -> None:
    ref = f"{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def fit_columns(ws, max_width: int = 36) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells[:400]:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def add_list_validation(ws, address: str, values: list[str]) -> None:
    quoted = ",".join(values)
    validation = DataValidation(type="list", formula1=f'"{quoted}"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(address)


def protect_support_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    ws.protection.sheet = True
    ws.protection.set_password(PROTECTION_PASSWORD)


def protect_entry_sheet(ws) -> None:
    editable_cols = {
        "A",  # Data Venda
        "B",  # Codigo Reserva
        "C",  # Hotel
        "D",  # Canal
        "E",  # Vendedor
        "F",  # Cliente
        "G",  # Checkin
        "H",  # Checkout
        "J",  # UHs
        "K",  # Adultos
        "L",  # Criancas
        "M",  # Valor Total
        "N",  # Recebido
        "P",  # Forma Pagto
        "Q",  # Parcelas
        "R",  # Status
        "T",  # Observacoes
    }
    for row in ws.iter_rows():
        for cell in row:
            col = get_column_letter(cell.column)
            is_input = cell.row > 1 and col in editable_cols
            cell.protection = Protection(locked=not is_input)
            if cell.row > 1:
                cell.fill = PatternFill("solid", fgColor=COLORS["input"] if is_input else COLORS["locked"])
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.set_password(PROTECTION_PASSWORD)


def add_catalog(wb: Workbook) -> None:
    ws = wb.create_sheet("Cadastros")
    sections = {
        "A": ("Hoteis", ["SUEDS PLAZA", "SUEDS SEGUNDO SOL", "SUEDS PREMIUM", "SUEDS CABRALIA", "SUEDS TRANCOSO"]),
        "C": ("Canais", ["BOOKING ENGINE", "BE MOBILE", "CENTRAL DE RESERVAS", "PARTICULAR", "Asksuite", "Google Ads", "Meta Ads", "Site", "Telefone", "Walk-in", "Operadoras", "OTAs", "Agencia", "Recepcao", "Outro"]),
        "E": ("Vendedores", ["Aline Nunes", "Amanda Melgaco", "Julia Reche", "Emanoel Cesar", "Site"]),
        "G": ("Pagamentos", ["PIX", "Cartao credito", "Cartao debito", "Dinheiro", "Boleto", "Transferencia", "Outros"]),
        "I": ("Status", ["Confirmada", "Pendente", "Cancelada", "No-show"]),
    }
    for col, (title, values) in sections.items():
        ws[f"{col}1"] = title
        ws[f"{col}1"].fill = PatternFill("solid", fgColor=COLORS["header"])
        ws[f"{col}1"].font = Font(color=COLORS["text"], bold=True)
        for idx, value in enumerate(values, start=2):
            ws[f"{col}{idx}"] = value
    ws.freeze_panes = "A2"
    fit_columns(ws)


def add_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("00_Instrucoes", 0)
    ws["A1"] = "SUEDS Hotels - Lançamento de Vendas Offline"
    ws["A1"].font = Font(size=18, bold=True, color=COLORS["text"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["dark"])
    ws.merge_cells("A1:F1")
    rows = [
        ["1", "Cada vendedor deve preencher somente a aba Lancamento_Vendas."],
        ["2", "Use uma linha por reserva. Nao agrupe varias reservas na mesma linha."],
        ["3", "Campos obrigatorios: Data Venda, Hotel, Canal, Vendedor, Cliente, Checkin, Checkout, Valor Total, Recebido, Forma Pagto e Status."],
        ["4", "A coluna A Receber calcula automaticamente Valor Total menos Recebido."],
        ["5", "Nao altere as abas Base_Dashboard, Dashboard_Base, Cadastros, De_Para e Funil_Asksuite sem alinhamento."],
        ["6", "Quando subir para Google Sheets, proteja as abas de apoio e deixe editavel apenas Lancamento_Vendas."],
    ]
    for r, row in enumerate(rows, start=3):
        ws.cell(r, 1).value = row[0]
        ws.cell(r, 2).value = row[1]
        ws.cell(r, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 120


def build_rows_from_source() -> list[list]:
    src = load_workbook(INPUT, data_only=True)
    ws = src["Vendas direta"]
    rows: list[list] = []
    seq = 1
    for row in ws.iter_rows(min_row=3, max_col=12, values_only=True):
        codigo, data_venda, hotel, canal, cliente, checkin, checkout, valor, recebido, a_receber, pagamento, vendedor = row
        if not any([codigo, data_venda, hotel, canal, cliente, valor]):
            continue
        forma, parcelas = payment_fields(str(pagamento or ""))
        canal_lancamento = canal or ""
        rows.append([
            data_venda,
            codigo,
            normalize_hotel(hotel),
            canal_lancamento,
            vendedor.title() if isinstance(vendedor, str) and vendedor.upper() != "SITE" else vendedor,
            cliente,
            checkin,
            checkout,
            f'=IF(OR(G{seq+1}="",H{seq+1}=""),"",H{seq+1}-G{seq+1})',
            1,
            "",
            "",
            valor or 0,
            recebido or 0,
            f'=IF(M{seq+1}="","",M{seq+1}-N{seq+1})',
            forma,
            parcelas,
            "Confirmada",
            "Historico V19jun26",
            str(pagamento or "") if forma in {"Outros", "Ver linha anterior"} else "",
        ])
        seq += 1
    return rows


def add_entry_sheet(wb: Workbook, rows: list[list]) -> None:
    ws = wb.create_sheet("Lancamento_Vendas")
    ws.append(ENTRY_HEADERS)
    for row in rows:
        ws.append(row)

    for idx in range(len(rows) + 1, 501):
        excel_row = idx + 1
        ws.append([
            "",
            "",
            "",
            "",
            "",
            "",
            f'=IF(OR(G{excel_row}="",H{excel_row}=""),"",H{excel_row}-G{excel_row})',
            1,
            "",
            "",
            "",
            "",
            f'=IF(M{excel_row}="","",M{excel_row}-N{excel_row})',
            "",
            "",
            "Confirmada",
            "Lancamento manual",
            "",
        ])

    style_header(ws, 1, 1, len(ENTRY_HEADERS))
    ws.freeze_panes = "A2"
    add_table(ws, "tbLancamentoVendas", 1, 1, ws.max_row, ws.max_column)
    add_list_validation(ws, "C2:C501", ["SUEDS PLAZA", "SUEDS SEGUNDO SOL", "SUEDS PREMIUM", "SUEDS CABRALIA", "SUEDS TRANCOSO"])
    add_list_validation(ws, "D2:D501", ["BOOKING ENGINE", "BE MOBILE", "CENTRAL DE RESERVAS", "PARTICULAR", "Asksuite", "Google Ads", "Meta Ads", "Site", "Telefone", "Walk-in", "Operadoras", "OTAs", "Agencia", "Recepcao", "Outro"])
    add_list_validation(ws, "E2:E501", ["Aline Nunes", "Amanda Melgaco", "Julia Reche", "Emanoel Cesar", "Site"])
    add_list_validation(ws, "P2:P501", ["PIX", "Cartao credito", "Cartao debito", "Dinheiro", "Boleto", "Transferencia", "Outros"])
    add_list_validation(ws, "R2:R501", ["Confirmada", "Pendente", "Cancelada", "No-show"])
    for col in ["A", "G", "H"]:
        for cell in ws[col][1:]:
            cell.number_format = "dd/mm/yyyy"
    for col in ["M", "N", "O"]:
        for cell in ws[col][1:]:
            cell.number_format = '"R$" #,##0.00'
    for col in ["I", "O", "S"]:
        ws.column_dimensions[col].hidden = col in {"S"}
    fit_columns(ws)


def add_base_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Base_Dashboard")
    ws.append(BASE_HEADERS)
    for i in range(2, 502):
        formulas = [
            f"=Lancamento_Vendas!A{i}",
            f"=Lancamento_Vendas!B{i}",
            f"=Lancamento_Vendas!C{i}",
            f"=Lancamento_Vendas!D{i}",
            f"=Lancamento_Vendas!E{i}",
            f"=Lancamento_Vendas!F{i}",
            f"=Lancamento_Vendas!G{i}",
            f"=Lancamento_Vendas!H{i}",
            f"=Lancamento_Vendas!I{i}",
            f"=Lancamento_Vendas!J{i}",
            f"=Lancamento_Vendas!K{i}",
            f"=Lancamento_Vendas!L{i}",
            f"=Lancamento_Vendas!M{i}",
            f"=Lancamento_Vendas!N{i}",
            f"=Lancamento_Vendas!O{i}",
            f"=Lancamento_Vendas!P{i}",
            f"=Lancamento_Vendas!Q{i}",
            f"=Lancamento_Vendas!R{i}",
            f"=Lancamento_Vendas!S{i}",
            f"=Lancamento_Vendas!T{i}",
            f"=Lancamento_Vendas!C{i}",
            f'=IF(D{i}="Operadoras","Operadora",IF(D{i}="OTAs","OTA","Direto"))',
            f'=IF(OR(D{i}="BOOKING ENGINE",D{i}="BE MOBILE"),"Booking Engine",IF(D{i}="PARTICULAR","WhatsApp",IF(D{i}="CENTRAL DE RESERVAS","Central de Reservas",D{i})))',
            f'=IF(E{i}="Site","Canal digital","Atendente")',
            f'=IF(A{i}="","",TEXT(A{i},"yyyy-mm"))',
        ]
        ws.append(formulas)
    style_header(ws, 1, 1, len(BASE_HEADERS), COLORS["dark"])
    ws.freeze_panes = "A2"
    add_table(ws, "tbBaseDashboard", 1, 1, ws.max_row, ws.max_column)
    fit_columns(ws)


def add_meta_funil_depara(wb: Workbook) -> None:
    metas = wb.create_sheet("Metas")
    metas.append(["Mes", "Tipo Meta", "Hotel", "Canal", "Responsavel", "Meta Receita", "Meta Reservas", "Observacoes"])
    for name in ["Aline Nunes", "Amanda Melgaco", "Julia Reche", "Emanoel Cesar", "Site", "Operadoras", "OTAs"]:
        metas.append(["2026-06", "Responsavel", "", "", name, "", "", "Preencher"])
    style_header(metas, 1, 1, 8)
    add_table(metas, "tbMetas", 1, 1, metas.max_row, metas.max_column)
    fit_columns(metas)

    depara = wb.create_sheet("De_Para")
    rows = [
        ["Tipo", "Valor Original", "Valor Padronizado", "Observacao"],
        ["Hotel", "SEGUNDO SOL", "SUEDS SEGUNDO SOL", ""],
        ["Hotel", "SEGUNDO CABRALIA", "SUEDS CABRALIA", "Confirmar regra"],
        ["Canal", "BE MOBILE", "Booking Engine", ""],
        ["Canal", "BOOKING ENGINE", "Booking Engine", ""],
        ["Canal", "PARTICULAR", "WhatsApp", "Confirmar se sempre deve ser WhatsApp"],
        ["Vendedor", "SITE", "Site", "Canal digital"],
    ]
    for row in rows:
        depara.append(row)
    style_header(depara, 1, 1, 4)
    add_table(depara, "tbDePara", 1, 1, depara.max_row, depara.max_column)
    fit_columns(depara)

    funil = wb.create_sheet("Funil_Asksuite")
    headers = ["Data", "Hotel/Robo", "Iniciou Atendimento", "Cotacao", "Reservar Agora", "Solicitou Humano", "Grupos", "Indisponibilidade", "Reservou", "Conversao Inicio>Reserva", "Observacoes"]
    funil.append(headers)
    for i in range(2, 63):
        funil.append(["", "Todos", "", "", "", "", "", "", "", f'=IF(C{i}=0,"",I{i}/C{i})', ""])
    style_header(funil, 1, 1, len(headers))
    add_table(funil, "tbFunilAsksuite", 1, 1, funil.max_row, funil.max_column)
    fit_columns(funil)


def add_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard_Base", 2)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Base do Dashboard - Vendas Offline"
    ws["A1"].font = Font(size=18, bold=True, color=COLORS["text"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["dark"])
    ws.merge_cells("A1:H1")

    rows = [
        ["Indicador", "Valor"],
        ["Receita Confirmada", '=SUMIFS(Base_Dashboard!M:M,Base_Dashboard!R:R,"Confirmada")'],
        ["Recebido", '=SUMIFS(Base_Dashboard!N:N,Base_Dashboard!R:R,"Confirmada")'],
        ["A Receber", '=SUMIFS(Base_Dashboard!O:O,Base_Dashboard!R:R,"Confirmada")'],
        ["Reservas Confirmadas", '=COUNTIFS(Base_Dashboard!R:R,"Confirmada",Base_Dashboard!A:A,"<>")'],
        ["Ticket Medio", '=IF(B6=0,0,B3/B6)'],
    ]
    for r, row in enumerate(rows, start=3):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value
    style_header(ws, 3, 1, 2, COLORS["header"])
    for cell in ws["B"][4:8]:
        cell.number_format = '"R$" #,##0.00'

    ws["D3"] = "Por Canal"
    ws["D3"].font = Font(bold=True, color=COLORS["text"])
    ws["D3"].fill = PatternFill("solid", fgColor=COLORS["header"])
    ws["D4"], ws["E4"] = "Canal", "Receita"
    channels = ["Central de Reservas", "WhatsApp", "Booking Engine", "Site", "Operadoras", "OTAs"]
    for idx, channel in enumerate(channels, start=5):
        ws.cell(idx, 4).value = channel
        ws.cell(idx, 5).value = f'=SUMIFS(Base_Dashboard!M:M,Base_Dashboard!W:W,D{idx},Base_Dashboard!R:R,"Confirmada")'
        ws.cell(idx, 5).number_format = '"R$" #,##0.00'
    style_header(ws, 4, 4, 5, COLORS["header"])

    ws["G3"] = "Por Vendedor"
    ws["G3"].font = Font(bold=True, color=COLORS["text"])
    ws["G3"].fill = PatternFill("solid", fgColor=COLORS["header"])
    ws["G4"], ws["H4"] = "Vendedor", "Receita"
    sellers = ["Aline Nunes", "Amanda Melgaco", "Julia Reche", "Emanoel Cesar", "Site"]
    for idx, seller in enumerate(sellers, start=5):
        ws.cell(idx, 7).value = seller
        ws.cell(idx, 8).value = f'=SUMIFS(Base_Dashboard!M:M,Base_Dashboard!E:E,G{idx},Base_Dashboard!R:R,"Confirmada")'
        ws.cell(idx, 8).number_format = '"R$" #,##0.00'
    style_header(ws, 4, 7, 8, COLORS["header"])

    chart = BarChart()
    chart.title = "Receita por Canal"
    chart.add_data(Reference(ws, min_col=5, min_row=4, max_row=10), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=4, min_row=5, max_row=10))
    chart.height = 7
    chart.width = 13
    ws.add_chart(chart, "D13")
    fit_columns(ws)


def main() -> None:
    rows = build_rows_from_source()
    wb = Workbook()
    wb.remove(wb.active)
    add_instructions(wb)
    add_entry_sheet(wb, rows)
    add_dashboard(wb)
    add_base_dashboard(wb)
    add_catalog(wb)
    add_meta_funil_depara(wb)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)

    for ws in wb.worksheets:
        if ws.title == "Lancamento_Vendas":
            protect_entry_sheet(ws)
        else:
            protect_support_sheet(ws)

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
