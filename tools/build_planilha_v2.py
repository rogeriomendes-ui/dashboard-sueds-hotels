from __future__ import annotations

import re
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\roger\OneDrive\Documentos\Dashboard SUEDS HOTELS")
INPUT = ROOT / "Vendas Sueds Hotels - V19jun26.xlsx"
OUTPUT = ROOT / "Vendas Sueds Hotels - V2 Dashboard.xlsx"


COLORS = {
    "dark": "11110F",
    "panel": "1B1D1B",
    "green": "38D39F",
    "blue": "4AA3FF",
    "yellow": "FFD166",
    "red": "FF6B6B",
    "text": "F5F7FB",
    "muted": "DDE3EA",
    "line": "D9E2EC",
}


BASE_HEADERS = [
    "Data Venda",
    "Codigo Reserva",
    "Hotel",
    "Hotel Normalizado",
    "Canal Original",
    "Canal Macro",
    "Canal Detalhado",
    "Origem",
    "Responsavel",
    "Tipo Responsavel",
    "Robo",
    "Campanha",
    "Cliente",
    "Checkin",
    "Checkout",
    "Diarias",
    "UHs",
    "Adultos",
    "Criancas",
    "Valor Total",
    "Recebido",
    "A Receber",
    "Forma Pagto Original",
    "Forma Pagto Padronizada",
    "Parcelas",
    "Status",
    "Fonte",
    "Observacoes",
]


def normalize_hotel(value: str) -> str:
    mapping = {
        "SEGUNDO SOL": "SUEDS SEGUNDO SOL",
        "SEGUNDO CABRALIA": "SUEDS CABRALIA",
        "SUEDS CABRALIA": "SUEDS CABRALIA",
        "SUEDS TRANCOSO": "SUEDS TRANCOSO",
        "SUEDS PLAZA": "SUEDS PLAZA",
        "SUEDS PREMIUM": "SUEDS PREMIUM",
    }
    return mapping.get((value or "").strip().upper(), value or "")


def channel_fields(channel: str, seller: str) -> tuple[str, str, str]:
    channel_key = (channel or "").strip().upper()
    seller_key = (seller or "").strip().upper()

    if channel_key in {"BOOKING ENGINE", "BE MOBILE"}:
        return "Direto", "Booking Engine", "Site"
    if channel_key == "CENTRAL DE RESERVAS":
        return "Direto", "Central de Reservas", "Central"
    if channel_key == "PARTICULAR":
        if seller_key == "SITE":
            return "Direto", "Site", "Site"
        return "Direto", "WhatsApp / Particular", "Central"
    return "Direto", channel or "Nao informado", "Nao informado"


def payment_fields(payment: str) -> tuple[str, str]:
    text = (payment or "").upper()
    installments = ""
    match = re.search(r"(\d+)\s*X", text)
    if match:
        installments = match.group(1)

    if "PIX" in text:
        return "PIX", installments
    if "CREDITO" in text or "CRÉDITO" in text or "CARTÃO" in text or "CARTAO" in text or "LINK" in text or "REDE" in text:
        return "Cartao credito", installments
    if "RECEP" in text:
        return "Recepcao", installments
    if "VALOR MENCIONADO" in text:
        return "Ver linha anterior", installments
    return "Outros / Manual", installments


def copy_original_sheet(src_ws, dst_wb: Workbook, title: str) -> None:
    ws = dst_wb.create_sheet(title)
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = ws[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
    for key, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[key].width = dim.width
    for key, dim in src_ws.row_dimensions.items():
        ws.row_dimensions[key].height = dim.height


def style_header(ws, row: int, start_col: int, end_col: int, fill: str = COLORS["panel"]) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=COLORS["text"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=COLORS["line"]))


def fit_columns(ws, max_width: int = 34) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells[:300]:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def add_table(ws, name: str, first_row: int, first_col: int, last_row: int, last_col: int) -> None:
    ref = f"{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def main() -> None:
    src = load_workbook(INPUT, data_only=False)
    src_values = load_workbook(INPUT, data_only=True)

    wb = Workbook()
    wb.remove(wb.active)

    copy_original_sheet(src["Vendas direta"], wb, "Original_Vendas_Direta")
    copy_original_sheet(src["Metas equipe"], wb, "Original_Metas_Equipe")

    readme = wb.create_sheet("00_Leia_me", 0)
    readme["A1"] = "Planilha base para Dashboard SUEDS Hotels"
    readme["A1"].font = Font(size=18, bold=True, color=COLORS["text"])
    readme["A1"].fill = PatternFill("solid", fgColor=COLORS["dark"])
    readme.merge_cells("A1:F1")
    notes = [
        ["Objetivo", "Transformar o registro de vendas em base analitica para dashboard em TV e celular."],
        ["Aba Base_Vendas", "Uma linha por reserva, com campos normalizados e campos futuros para BI."],
        ["Aba Metas", "Metas por mes, hotel, canal e responsavel."],
        ["Aba Funil_Asksuite", "Entrada diaria das etapas do funil ate a integracao automatica."],
        ["Aba De_Para", "Padronizacao de hoteis, canais, vendedores e pagamentos."],
        ["Observacao", "As abas Original_* preservam o arquivo recebido como referencia."],
    ]
    for r, row in enumerate(notes, start=3):
        readme.cell(r, 1).value = row[0]
        readme.cell(r, 2).value = row[1]
        readme.cell(r, 1).font = Font(bold=True)
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 96

    base = wb.create_sheet("Base_Vendas")
    base.append(BASE_HEADERS)

    source_ws = src_values["Vendas direta"]
    for row in source_ws.iter_rows(min_row=3, max_col=12, values_only=True):
        codigo, data_venda, hotel, canal, cliente, checkin, checkout, valor, recebido, a_receber, pagamento, vendedor = row
        if not any([codigo, data_venda, hotel, canal, cliente, valor]):
            continue
        hotel_norm = normalize_hotel(hotel)
        canal_macro, canal_detalhado, origem = channel_fields(canal, vendedor)
        tipo_responsavel = "Canal digital" if (vendedor or "").strip().upper() == "SITE" else "Atendente"
        forma_padrao, parcelas = payment_fields(pagamento)
        diarias = ""
        if isinstance(checkin, datetime) and isinstance(checkout, datetime):
            diarias = (checkout - checkin).days
        base.append([
            data_venda,
            codigo,
            hotel,
            hotel_norm,
            canal,
            canal_macro,
            canal_detalhado,
            origem,
            vendedor,
            tipo_responsavel,
            "",
            "",
            cliente,
            checkin,
            checkout,
            diarias,
            1,
            "",
            "",
            valor or 0,
            recebido or 0,
            a_receber or 0,
            pagamento,
            forma_padrao,
            parcelas,
            "Confirmada",
            "Planilha Vendas Diretas",
            "",
        ])

    style_header(base, 1, 1, len(BASE_HEADERS), COLORS["dark"])
    base.freeze_panes = "A2"
    add_table(base, "tbBaseVendas", 1, 1, base.max_row, base.max_column)
    for col in ["A", "N", "O"]:
        for cell in base[col][1:]:
            cell.number_format = "dd/mm/yyyy"
    for col in ["T", "U", "V"]:
        for cell in base[col][1:]:
            cell.number_format = '"R$" #,##0.00'
    fit_columns(base)

    metas = wb.create_sheet("Metas")
    metas_headers = ["Mes", "Tipo Meta", "Hotel", "Canal Detalhado", "Responsavel", "Meta Receita", "Meta Reservas", "Observacoes"]
    metas.append(metas_headers)
    for responsavel in ["Aline Nunes", "Amanda Melgaco", "Julia Reche", "Emanoel Cesar", "Site", "Operadoras", "OTAs", "BE Mobile"]:
        metas.append(["2026-05", "Responsavel", "", "", responsavel, "", "", "Preencher meta"])
    style_header(metas, 1, 1, len(metas_headers), COLORS["dark"])
    add_table(metas, "tbMetas", 1, 1, metas.max_row, metas.max_column)
    fit_columns(metas)

    depara = wb.create_sheet("De_Para")
    depara_rows = [
        ["Tipo", "Valor Original", "Valor Padronizado", "Observacao"],
        ["Hotel", "SEGUNDO SOL", "SUEDS SEGUNDO SOL", ""],
        ["Hotel", "SEGUNDO CABRALIA", "SUEDS CABRALIA", "Possivel divergencia no lancamento original"],
        ["Canal", "BOOKING ENGINE", "Booking Engine", ""],
        ["Canal", "BE MOBILE", "Booking Engine", ""],
        ["Canal", "CENTRAL DE RESERVAS", "Central de Reservas", ""],
        ["Canal", "PARTICULAR", "WhatsApp / Particular", ""],
        ["Vendedor", "SITE", "Site", "Tratar como canal digital, nao atendente"],
        ["Pagamento", "CARTAO/CARTÃO CREDITO", "Cartao credito", ""],
        ["Pagamento", "PIX", "PIX", ""],
    ]
    for row in depara_rows:
        depara.append(row)
    style_header(depara, 1, 1, 4, COLORS["dark"])
    add_table(depara, "tbDePara", 1, 1, depara.max_row, depara.max_column)
    fit_columns(depara)

    funil = wb.create_sheet("Funil_Asksuite")
    funil_headers = ["Data", "Hotel/Robo", "Iniciou Atendimento", "Cotacao", "Reservar Agora", "Solicitou Humano", "Grupos", "Indisponibilidade", "Reservou", "Conversao Inicio>Reserva", "Observacoes"]
    funil.append(funil_headers)
    for i in range(1, 32):
        funil.append([f"2026-05-{i:02d}", "Todos", "", "", "", "", "", "", "", f'=IF(C{i+1}=0,"",I{i+1}/C{i+1})', ""])
    style_header(funil, 1, 1, len(funil_headers), COLORS["dark"])
    add_table(funil, "tbFunilAsksuite", 1, 1, funil.max_row, funil.max_column)
    for cell in funil["J"][1:]:
        cell.number_format = "0.0%"
    fit_columns(funil)

    dash = wb.create_sheet("Dashboard_Base", 2)
    dash.sheet_view.showGridLines = False
    dash["A1"] = "Resumo para Dashboard"
    dash["A1"].font = Font(size=18, bold=True, color=COLORS["text"])
    dash["A1"].fill = PatternFill("solid", fgColor=COLORS["dark"])
    dash.merge_cells("A1:H1")

    kpis = [
        ["Indicador", "Valor"],
        ["Receita Total", "=SUM(Base_Vendas!T:T)"],
        ["Recebido", "=SUM(Base_Vendas!U:U)"],
        ["A Receber", "=SUM(Base_Vendas!V:V)"],
        ["Reservas", "=COUNTA(Base_Vendas!B:B)-1"],
        ["Ticket Medio", "=IF(B5=0,0,B2/B5)"],
    ]
    for r, row in enumerate(kpis, start=3):
        for c, value in enumerate(row, start=1):
            dash.cell(r, c).value = value
    style_header(dash, 3, 1, 2, COLORS["panel"])
    for cell in dash["B"][3:8]:
        cell.number_format = '"R$" #,##0.00'
    dash["B7"].number_format = '"R$" #,##0.00'

    dash["D3"] = "Vendas por Canal"
    dash["D3"].font = Font(bold=True, color=COLORS["text"])
    dash["D3"].fill = PatternFill("solid", fgColor=COLORS["panel"])
    dash["D4"], dash["E4"] = "Canal", "Receita"
    channels = ["Booking Engine", "Central de Reservas", "WhatsApp / Particular"]
    for idx, channel in enumerate(channels, start=5):
        dash.cell(idx, 4).value = channel
        dash.cell(idx, 5).value = f'=SUMIF(Base_Vendas!G:G,D{idx},Base_Vendas!T:T)'
        dash.cell(idx, 5).number_format = '"R$" #,##0.00'
    style_header(dash, 4, 4, 5, COLORS["panel"])

    dash["G3"] = "Vendas por Hotel"
    dash["G3"].font = Font(bold=True, color=COLORS["text"])
    dash["G3"].fill = PatternFill("solid", fgColor=COLORS["panel"])
    dash["G4"], dash["H4"] = "Hotel", "Receita"
    hotels = ["SUEDS PLAZA", "SUEDS SEGUNDO SOL", "SUEDS TRANCOSO", "SUEDS CABRALIA", "SUEDS PREMIUM"]
    for idx, hotel in enumerate(hotels, start=5):
        dash.cell(idx, 7).value = hotel
        dash.cell(idx, 8).value = f'=SUMIF(Base_Vendas!D:D,G{idx},Base_Vendas!T:T)'
        dash.cell(idx, 8).number_format = '"R$" #,##0.00'
    style_header(dash, 4, 7, 8, COLORS["panel"])

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Receita por Hotel"
    chart.y_axis.title = "Hotel"
    chart.x_axis.title = "Receita"
    chart.add_data(Reference(dash, min_col=8, min_row=4, max_row=9), titles_from_data=True)
    chart.set_categories(Reference(dash, min_col=7, min_row=5, max_row=9))
    chart.height = 7
    chart.width = 13
    dash.add_chart(chart, "G12")
    fit_columns(dash)

    status_validation = DataValidation(type="list", formula1='"Confirmada,Cancelada,Pendente,No-show"', allow_blank=False)
    base.add_data_validation(status_validation)
    status_validation.add(f"Z2:Z{base.max_row + 500}")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = copy(cell.alignment)
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "left",
                    vertical="center",
                    wrap_text=cell.alignment.wrap_text,
                )

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
